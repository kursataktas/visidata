
from visidata import *

class open_xlsx(Sheet):
    'Load XLSX file (in Excel Open XML format).'
    commands = [
        Command(ENTER, 'vd.push(sheet.getSheet(cursorRow))', 'loads the entire table into memory')
    ]
    def __init__(self, path):
        super().__init__(path.name, path)
        self.workbook = None

    @async
    def reload(self):
        import openpyxl
        self.columns = [Column('name')]
        self.workbook = openpyxl.load_workbook(self.source.resolve(), data_only=True, read_only=True)
        self.rows = list(self.workbook.sheetnames)

    def getSheet(self, sheetname):
        worksheet = self.workbook.get_sheet_by_name(sheetname)
        return xlsxSheet(joinSheetnames(self.name, sheetname), worksheet)

class xlsxSheet(Sheet):
    @async
    def reload(self):
        worksheet = self.source
        self.columns = ArrayColumns(worksheet.max_column)
        with Progress(self, worksheet.max_row) as prog:
            for row in worksheet.iter_rows():
                self.addRow([cell.value for cell in row])
                self.addProgress(1)

class open_xls(Sheet):
    'Load XLS file (in Excel format).'
    commands = [
        Command(ENTER, 'vd.push(sheet.getSheet(cursorRow))', 'loads the entire table into memory')
    ]
    def __init__(self, path):
        super().__init__(path.name, path)
        self.workbook = None

    @async
    def reload(self):
        import xlrd
        self.columns = [Column('name')]
        self.workbook = xlrd.open_workbook(self.source.resolve())
        self.rows = list(self.workbook.sheet_names())

    def getSheet(self, sheetname):
        worksheet = self.workbook.sheet_by_name(sheetname)
        return xlsSheet(joinSheetnames(self.name, sheetname), worksheet)


class xlsSheet(Sheet):
    @async
    def reload(self):
        worksheet = self.source
        self.columns = ArrayColumns(worksheet.ncols)
        with Progress(self, worksheet.nrows) as prog:
            for rownum in range(worksheet.nrows):
                self.addRow([worksheet.cell(rownum, colnum).value for colnum in range(worksheet.ncols)])
                prog.addProgress(1)
